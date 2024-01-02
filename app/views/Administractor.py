import openpyxl
from django.views import View
from rest_framework.views import APIView, Request
from rest_framework.response import Response

from app.models import Order, User, Administrator
from app.utils import get_header_token, decode_token
from app.views.Util import judgeAdministrator


class GetIfAdmin(APIView):
    def get(self, request):
        token = get_header_token(request)
        ifAdmin = False
        if not decode_token(token):
            code, message = -1, '登录超时或者其他原因导致token失效'
        else:
            userName = decode_token(token)['username']
            # ifAdmin = (userName == 'Admin')
            ifAdmin = judgeAdministrator(userName)
            code, message = 200, '成功查询是否是管理员'
        return Response({'code': code, 'message': message, 'data': {'ifAdmin': ifAdmin}})


class GetAllUsers(APIView):
    def get(self, request):
        token = get_header_token(request)
        data = []
        if not decode_token(token):
            code, message = -1, '登录超时或者其他原因导致token失效'
        else:
            userName = decode_token(token)['username']

            if not judgeAdministrator(userName):
                code, message = -2, '您不是管理员，无权限进行此操作'
            else:
                user_list_raw = User.objects.all()
                for u in user_list_raw:
                    if judgeAdministrator(u.name):
                        continue
                    data.append({
                        'userName': u.name,
                        'avatar': None if not u.avatar else u.avatar.get_url(),
                        'email': u.email,
                        'registerTime': u.get_register_time()
                    })
                code, message = 200, '获取所有普通用户成功'
        return Response({'code': code, 'message': message, 'data': data})


class GetAllOrders(APIView):
    def get(self, request):
        token = get_header_token(request)
        data = []
        if not decode_token(token):
            code, message = -1, '登录超时或者其他原因导致token失效'
        else:
            userName = decode_token(token)['username']
            if not judgeAdministrator(userName):
                code, message = -2, '您不是管理员，无权限进行此操作'
            else:
                order_list_raw = Order.objects.all().order_by('-create_time')
                for o in order_list_raw:
                    data.append({
                        'id': o.id,
                        'buyer': {
                            'userName': o.buyer_name,
                            # 'avatar': None if not o.buyer.avatar else o.buyer.avatar.get_url(),
                            # 'email': o.buyer.email
                        },
                        'seller': {
                            'userName': o.seller_name,
                            # 'avatar': None if not o.buyer.avatar else o.buyer.avatar.get_url(),
                            # 'email': o.buyer.email
                        },
                        'product': {
                            'productName': o.product_name,
                            'price': o.price,
                        },
                        'number': o.number,
                        'status': o.status,
                        'createdTime': o.get_create_time(),
                        'buyerInfo': {
                            'name': o.receiver_name,
                            'phone': o.receiver_phone,
                            'place': o.receiver_place
                        }
                    })
                code, message = 200, '获取所有订单成功'
        return Response({'code': code, 'message': message, 'data': data})


class DeleteUser(APIView):
    def post(self, request):
        token = get_header_token(request)
        data = []
        if not decode_token(token):
            code, message = -1, '登录超时或者其他原因导致token失效'
        else:
            userName = decode_token(token)['username']
            if not judgeAdministrator(userName):
                code, message = -2, '您不是管理员，无权限进行此操作'
            else:
                delete_userName = request.data.get('userName')
                try:
                    u = User.objects.get(name=delete_userName)
                    if judgeAdministrator(u.name) and userName != 'Admin':
                        code, message = -3, '您没有删除管理员的权限'
                    else:
                        u.delete()
                        print()
                        code, message = 200, '用户删除成功'
                except User.DoesNotExist:
                    code, message = -2, '要删除的用户不存在'
        return Response({'code': code, 'message': message})


class DeleteOrder(APIView):
    def post(self, request):
        token = get_header_token(request)
        if not decode_token(token):
            code, message = -1, '登录超时或者其他原因导致token失效'
        else:
            userName = decode_token(token)['username']
            if not judgeAdministrator(userName):
                code, message = -2, '您不是管理员，无权限进行此操作'
            else:
                delete_order_id = request.data.get('id')
                try:
                    o = Order.objects.get(id=delete_order_id)
                    o.delete()
                    code, message = 200, '订单删除成功'
                except Order.DoesNotExist:
                    code, message = -2, '要删除的订单不存在'
        return Response({'code': code, 'message': message})


class UploadUsers(APIView):
    def post(self, request):
        code, message = 200, '批量添加操作完成 '
        file = request.FILES.get("file")
        if not file:
            code, message = -1, '未上传文件'
        # 处理Excel文件
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        # 检查表头
        b1_cell = (sheet.cell(row=1, column=2)).value
        c1_cell = (sheet.cell(row=1, column=3)).value
        d1_cell = (sheet.cell(row=1, column=4)).value
        if b1_cell != '用户名' or c1_cell != '密码' or d1_cell != '邮箱':
            code, message = -2, '添加失败，表格格式错误'
        else:
            for row in sheet.iter_rows(min_row=2):
                if all(cell.value is None for cell in row):
                    continue
                name = str(row[1].value)
                email = str(row[3].value)
                password = str(row[2].value)
                list = User.objects.filter(name=name)
                if list:
                    code = -3
                    message = message + '用户' + str(name) + '添加失败，已经有同名用户存在\n'
                    continue
                u = User.objects.create(name=name, email=email, password=password)
                u.save()
        return Response({'code': code, 'message': message})


class AddNewAdministrator(APIView):
    def post(self, request):
        token = get_header_token(request)
        if not decode_token(token):
            code, message = -1, '登录超时或者其他原因导致token失效'
        else:
            userName = decode_token(token)['username']
            if userName != 'Admin':
                code, message = -2, '您无权添加管理员'
            else:
                name = request.data.get('userName')
                try:
                    u = User.objects.get(name=name)
                    Administrator.objects.create(user=u, level=1)
                    code, message = 200, '为用户添加权限成功'
                except User.DoesNotExist:
                    code, message = -2, '添加权限的用户不存在'
        return Response({'code': code, 'message': message})


class AdministratorSelect(APIView):
    def get(self, request):
        data = []
        token = get_header_token(request)
        if not decode_token(token):
            code, message = -1, '登录超时或者其他原因导致token失效'
        else:
            userName = decode_token(token)['username']
            if userName != 'Admin':
                code, message = -2, '您无权查看其他管理员'
            else:
                admin_list_raw = Administrator.objects.all()
                for ad in admin_list_raw:
                    if ad.user.name == 'Admin':
                        continue
                    data.append({
                        'userName': ad.user.name,
                        'avatar': None if not ad.user.avatar else ad.user.avatar.get_url(),
                        'email': ad.user.email,
                        'registerTime': ad.user.get_register_time()
                    })
                code, message = 200, '获取所有管理员成功'
        return Response({'code': code, 'message': message, 'data': data})

